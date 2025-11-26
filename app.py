import streamlit as st
import pandas as pd
from datetime import date
import io
from openpyxl import Workbook
from openpyxl.styles import Font

st.set_page_config(
    page_title="Grow Tracker",
    page_icon="https://raw.githubusercontent.com/michaelreed452/cannabis-grow-tracker/main/leaf.png",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===================== EXACT-MATCH LOGIN  =====================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.display_name = ""

if not st.session_state.logged_in:
    st.markdown("### Login to Grow Tracker")

    with st.form("exact_login"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")

        # ←←← TYPE EXACTLY THESE IN THE BOXES TO LOG IN ←←←
        EXACT_USERS = {
            "Michael": "KATVIS",    # ← change ONLY the password if you want
            "Fanie":   "GhostOG420",
            # add more exactly like this
        }

        if st.form_submit_button("Login", type="primary"):
            if username in EXACT_USERS and EXACT_USERS[username] == password:
                st.session_state.logged_in = True
                st.session_state.display_name = username
                st.rerun()
            else:
                st.error("Wrong. Type exactly as written in the code.")

    st.stop()

st.sidebar.success(f"Logged in as **{st.session_state.display_name}**")

# ===================== DATA INIT =====================
def init():
    if "plants" not in st.session_state:
        st.session_state.plants = pd.DataFrame(columns=[
            "Plant ID","Strain Name","Variety","Gender","Environment","Type","Source","Batch #",
            "Date Germination","Date Transplant Veg","Date Flip Flower","Date Harvest",
            "Wet Weight (g)","Dry Weight (g)","Trimmed Yield (g)","Mother ID",
            "Growing Medium","Container Size (L)","Phenotype Notes","Health Issues",
            "Rating (1-10)","Photos Link","Status"
        ])
    if "strains" not in st.session_state:
        st.session_state.strains = pd.DataFrame(columns=["Strain Name","Breeder","Variety","Expected Flower Time","THC %","Terpene Profile","Average Yield (g/plant)","Times Grown","Best Pheno Notes","Keeper?"])
    if "expenses" not in st.session_state:
        st.session_state.expenses = pd.DataFrame(columns=["Date","Category","Item","Supplier","Cost (ZAR)","Quantity","Paid To","Notes","Receipt Link"])
    if "income" not in st.session_state:
        st.session_state.income = pd.DataFrame(columns=["Date","Strain","Grams Sold","Price per Gram","Buyer/Channel","Payment Method","Notes"])
    if "stock" not in st.session_state:
        st.session_state.stock = pd.DataFrame(columns=["Strain","Breeder","Seeds/Clones Left","Pack Cost (ZAR)"])
    if "feeding" not in st.session_state:
        st.session_state.feeding = pd.DataFrame(columns=[
            "Date","Plant IDs","Stage","Nutrient 1","Amount 1","Nutrient 2","Amount 2",
            "Nutrient 3","Amount 3","Nutrient 4","Amount 4","Nutrient 5","Amount 5","Notes"
        ])
init()

NUTRIENTS = ["CalMag Essential","NC32","Pot Grow","Pot Flora","Pot Radix","Bio-Blend","Carbon K","Multi Foliar Spray Concentrate","Other"]

# Auto calculate current stage
def get_stage(row):
    today = date.today()
    germ = row["Date Germination"]
    veg = row["Date Transplant Veg"]
    flip = row["Date Flip Flower"]
    harvest = row["Date Harvest"]
    if pd.isna(germ): return "Not Started"
    if pd.notna(harvest) and harvest <= today: return "Harvested"
    if pd.notna(flip) and flip <= today:
        weeks = (today - flip).days // 7 + 1
        return f"Flower Week {weeks}"
    if pd.notna(veg) and veg <= today:
        weeks = (today - veg).days // 7 + 1
        return f"Veg Week {weeks}"
    weeks = (today - germ).days // 7 + 1
    return f"Seedling Week {weeks}"

if len(st.session_state.plants) > 0:
    st.session_state.plants["Current Stage"] = st.session_state.plants.apply(get_stage, axis=1)

# ===================== SIDEBAR =====================
st.sidebar.markdown("# Grow Tracker")
pages = ["Dashboard","Plants Tracker","Strains Library","Expenses","Income","Seed & Clone Stock","Feeding Schedule","Export to Excel"]
emojis = ["Home","Plants","Strains","Expenses","Income","Stock","Feeding","Excel"]

for i, p in enumerate(pages):
    if st.sidebar.button(f"{emojis[i]} {p}", use_container_width=True,
                        type="primary" if st.session_state.get("page") == p else "secondary"):
        st.session_state.page = p
page = st.session_state.get("page", "Dashboard")

# ===================== PAGES =====================
if page == "Dashboard":
    st.title("Home Dashboard")
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Plants", len(st.session_state.plants))
    c2.metric("Yield", f"{st.session_state.plants['Trimmed Yield (g)'].sum():.1f} g")
    c3.metric("Expenses", f"R {st.session_state.expenses['Cost (ZAR)'].sum():,.2f}")
    income = (st.session_state.income["Grams Sold"] * st.session_state.income["Price per Gram"]).sum() if len(st.session_state.income)>0 else 0
    c4.metric("Income", f"R {income:,.2f}")

elif page == "Plants Tracker":
    st.title("Plants Tracker")
    tab1, tab2 = st.tabs(["View & Edit", "Add New"])

    with tab1:
        if len(st.session_state.plants) > 0:
            edited = st.data_editor(
                st.session_state.plants,
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True
            )
            if st.button("Save Changes", type="primary"):
                st.session_state.plants = edited
                st.success("All plants updated!")
                st.rerun()
        else:
            st.info("No plants yet — add one in the next tab")

    with tab2:
        with st.form("add_plant_form"):
            st.subheader("Add New Plant")
            col1, col2 = st.columns(2)
            with col1:
                plant_id = st.text_input("Plant ID *")
                strain = st.text_input("Strain Name *")
                variety = st.selectbox("Variety", ["Sativa", "Indica", "Hybrid"])
                gender = st.selectbox("Gender", ["Female", "Male", "Unknown"])
                environment = st.selectbox("Environment", ["Indoor", "Outdoor", "Greenhouse"])
                medium = st.selectbox("Growing Medium", ["Fabric Pot","Plastic Pot","Direct Soil","Coco","Hydro","Air Pot"])
                container = st.number_input("Container Size (L)", min_value=0.0, step=0.5)
            with col2:
                source = st.text_input("Source / Batch #")
                germ_date = st.date_input("Date Germination", date.today())
                veg_date = st.date_input("Date Transplant Veg (optional)", value, value=None)
                flip_date = st.date_input("Date Flip to Flower (optional)", value=None)
                harvest_date = st.date_input("Date Harvest (optional)", value=None)
                notes = st.text_area("Phenotype / Health Notes")

            if st.form_submit_button("Add Plant", type="primary"):
                if not plant_id or not strain:
                    st.error("Plant ID and Strain Name are required")
                else:
                    new_plant = pd.DataFrame([{
                        "Plant ID": plant_id,
                        "Strain Name": strain,
                        "Variety": variety,
                        "Gender": gender,
                        "Environment": environment,
                        "Type": "",
                        "Source": source,
                        "Batch #": source,
                        "Date Germination": pd.to_datetime(germ_date),
                        "Date Transplant Veg": pd.to_datetime(veg_date) if veg_date else pd.NaT,
                        "Date Flip Flower": pd.to_datetime(flip_date) if flip_date else pd.NaT,
                        "Date Harvest": pd.to_datetime(harvest_date) if harvest_date else pd.NaT,
                        "Wet Weight (g)": 0.0,
                        "Dry Weight (g)": 0.0,
                        "Trimmed Yield (g)": 0.0,
                        "Mother ID": "",
                        "Growing Medium": medium,
                        "Container Size (L)": container,
                        "Phenotype Notes": notes,
                        "Health Issues": "",
                        "Rating (1-10)": 0,
                        "Photos Link": "",
                        "Status": "Germinating"
                    }])
                    st.session_state.plants = pd.concat([st.session_state.plants, new_plant], ignore_index=True)
                    st.success(f"Plant {plant_id} added!")
                    st.rerun()

elif page == "Feeding Schedule":
    st.title("Feeding Schedule")
    t1,t2 = st.tabs(["Add Feeding","History"])
    with t1:
        with st.form("feed"):
            date_feed = st.date_input("Date", date.today())
            if len(st.session_state.plants)==0:
                st.warning("No plants yet")
            else:
                mode = st.radio("Select", ["Single Plant","Group"])
                if mode == "Single Plant":
                    pid = st.selectbox("Plant ID", st.session_state.plants["Plant ID"])
                    plant_ids = [pid]
                    stage = st.session_state.plants.loc[st.session_state.plants["Plant ID"]==pid, "Current Stage"].iloc[0]
                else:
                    group = st.selectbox("Group", ["All Plants","All Veg","All Flower"] + sorted(st.session_state.plants["Current Stage"].dropna().unique()))
                    if group=="All Plants": plant_ids = st.session_state.plants["Plant ID"].tolist()
                    elif group=="All Veg": plant_ids = st.session_state.plants[st.session_state.plants["Current Stage"].str.contains("Veg|Seedling")]["Plant ID"].tolist()
                    elif group=="All Flower": plant_ids = st.session_state.plants[st.session_state.plants["Current Stage"].str.contains("Flower")]["Plant ID"].tolist()
                    else: plant_ids = st.session_state.plants[st.session_state.plants["Current Stage"]==group]["Plant ID"].tolist()
                    stage = group
                    st.info(f"Selected: {', '.join(plant_ids)}")

                nut1 = st.selectbox("Nutrient 1", NUTRIENTS)
                amt1 = st.number_input("Amount 1 (ml/g)", 0.0, step=0.1)
                if nut1=="Other": nut1 = st.text_input("Specify other nutrient")

                more = st.checkbox("Add more nutrients")
                extras = []
                if more:
                    count = st.slider("How many more?",1,4,1)
                    for i in range(count):
                        n = st.selectbox(f"Nutrient {i+2}", NUTRIENTS, key=f"n{i}")
                        a = st.number_input(f"Amount {i+2}", 0.0, step=0.1, key=f"a{i}")
                        if n=="Other": n = st.text_input("Specify", key=f"o{i}")
                        extras.append((n,a))

                notes = st.text_area("Notes")

                if st.form_submit_button("Save Feeding"):
                    row = {"Date":date_feed,"Plant IDs":", ".join(plant_ids),"Stage":stage,
                           "Nutrient 1":nut1,"Amount 1":amt1,"Notes":notes}
                    for i,(n,a) in enumerate(extras):
                        row[f"Nutrient {i+2}"] = n
                        row[f"Amount {i+2}"] = a
                    st.session_state.feeding = pd.concat([st.session_state.feeding, pd.DataFrame([row])], ignore_index=True)
                    st.success("Feeding recorded!")
                    st.rerun()

    with t2:
        if len(st.session_state.feeding)>0:
            st.dataframe(st.session_state.feeding.sort_values("Date", ascending=False), use_container_width=True, hide_index=True)
        else:
            st.info("No feedings yet")

# (Keep your other pages — Strains, Expenses, Income, Stock — exactly as before)

elif page == "Export to Excel":
    st.title("Export to Excel")
    if st.button("Generate File", type="primary"):
        wb = Workbook()
        wb.remove(wb.active)
        for name, df in [("Plants",st.session_state.plants),("Feeding",st.session_state.feeding),
                         ("Strains",st.session_state.strains),("Expenses",st.session_state.expenses),
                         ("Income",st.session_state.income),("Stock",st.session_state.stock)]:
            ws = wb.create_sheet(name[:31])
            for c,col in enumerate(df.columns,1):
                ws.cell(1,c,col).font = Font(bold=True)
            for r,row in enumerate(df.itertuples(index=False),2):
                for c,val in enumerate(row,1):
                    ws.cell(r,c,val)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        st.download_button("Download Excel", buffer, f"GrowTracker_{date.today()}.xlsx")

# ===================== FOOTER =====================
st.markdown("---")
c1,c2,c3,c4,c5 = st.columns(5)
with c1: st.markdown(f"<div style='text-align:center'>Plants<br><b>{len(st.session_state.plants)}</b></div>", unsafe_allow_html=True)
with c2: st.markdown(f"<div style='text-align:center'>Strains<br><b>{len(st.session_state.strains)}</b></div>", unsafe_allow_html=True)
with c3: st.markdown(f"<div style='text-align:center'>Expenses<br><b>{len(st.session_state.expenses)}</b></div>", unsafe_allow_html=True)
with c4: st.markdown(f"<div style='text-align:center'>Income<br><b>{len(st.session_state.income)}</b></div>", unsafe_allow_html=True)
with c5: st.markdown(f"<div style='text-align:center'>Stock<br><b>{len(st.session_state.stock)}</b></div>", unsafe_allow_html=True)
